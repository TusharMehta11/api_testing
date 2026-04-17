"""
Write API test results to a formatted .xlsx file.

Sheets produced
---------------
1. Summary   — aggregate counts + duration, one row per endpoint
2. Details   — one row per test case variant with full request/response data
3. Latency   — latency breakdown (TTFB / body / total) per test case
"""

from __future__ import annotations

import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
COLOUR = {
    "header_bg":  "1F3864",   # dark navy — sheet header row
    "header_fg":  "FFFFFF",
    "pass_bg":    "C6EFCE",   # light green
    "pass_fg":    "276221",
    "fail_bg":    "FFC7CE",   # light red
    "fail_fg":    "9C0006",
    "warn_bg":    "FFEB9C",   # light yellow
    "warn_fg":    "9C5700",
    "alt_row":    "EEF2FF",   # alternating row tint
    "title_bg":   "2E75B6",   # summary title bar
}

THIN = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _fill(hex_colour: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_colour)


def _font(bold: bool = False, colour: str = "000000", size: int = 10) -> Font:
    return Font(bold=bold, color=colour, size=size, name="Calibri")


def _header_cell(ws, row: int, col: int, value: str) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(bold=True, colour=COLOUR["header_fg"], size=10)
    cell.fill = _fill(COLOUR["header_bg"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def _verdict_style(verdict: str) -> tuple[str, str]:
    """Return (bg, fg) hex pair for a verdict string."""
    v = (verdict or "").upper()
    if v == "PASS":
        return COLOUR["pass_bg"], COLOUR["pass_fg"]
    if v == "FAIL":
        return COLOUR["fail_bg"], COLOUR["fail_fg"]
    if v == "WARN":
        return COLOUR["warn_bg"], COLOUR["warn_fg"]
    return "FFFFFF", "000000"


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

def _build_summary(wb: Workbook, results: list[dict], run_meta: dict) -> None:
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    # ---- Title bar ----
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "API Test Run — Summary"
    title_cell.font = _font(bold=True, colour="FFFFFF", size=14)
    title_cell.fill = _fill(COLOUR["title_bg"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ---- Run-level stats (A3:B7) ----
    stats = [
        ("Collection", run_meta.get("collection", "")),
        ("Run started", run_meta.get("started_at", "")),
        ("Run finished", run_meta.get("finished_at", "")),
        ("Total tests", run_meta.get("total", 0)),
        ("Passed", run_meta.get("passed", 0)),
        ("Failed", run_meta.get("failed", 0)),
        ("Warned", run_meta.get("warned", 0)),
        ("Errors", run_meta.get("errors", 0)),
    ]
    for r, (label, value) in enumerate(stats, start=3):
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = _font(bold=True)
        lc.fill = _fill("DCE6F1")
        lc.border = THIN_BORDER
        lc.alignment = Alignment(horizontal="left", vertical="center")

        vc = ws.cell(row=r, column=2, value=value)
        vc.font = _font()
        vc.border = THIN_BORDER
        vc.alignment = Alignment(horizontal="left", vertical="center")

    # ---- Per-endpoint table ----
    headers = [
        "Endpoint Name", "Folder", "Method", "URL",
        "Total Variants", "Passed", "Failed", "Warned",
        "Avg TTFB (ms)", "Avg Body (ms)", "Avg Total (ms)",
        "Min Total (ms)", "Max Total (ms)",
    ]
    header_row = 13
    for c, h in enumerate(headers, start=1):
        _header_cell(ws, header_row, c, h)
    ws.row_dimensions[header_row].height = 20

    # Aggregate per endpoint
    endpoint_stats: dict[str, dict] = {}
    for res in results:
        key = res.get("endpoint_name", res.get("url", ""))
        if key not in endpoint_stats:
            endpoint_stats[key] = {
                "folder": res.get("folder_path", ""),
                "method": res.get("method", ""),
                "url": res.get("url", ""),
                "total": 0, "passed": 0, "failed": 0, "warned": 0,
                "ttfb_list": [], "body_list": [], "total_list": [],
            }
        s = endpoint_stats[key]
        s["total"] += 1
        verdict = (res.get("analysis", {}) or {}).get("verdict", "FAIL").upper()
        if verdict == "PASS":
            s["passed"] += 1
        elif verdict == "WARN":
            s["warned"] += 1
        else:
            s["failed"] += 1
        if not res.get("error"):
            s["ttfb_list"].append(res.get("request_time_ms", 0.0))
            s["body_list"].append(res.get("response_time_ms", 0.0))
            s["total_list"].append(res.get("total_time_ms", 0.0))

    def _avg(lst: list) -> str:
        return f"{sum(lst)/len(lst):.1f}" if lst else "—"

    def _min(lst: list) -> str:
        return f"{min(lst):.1f}" if lst else "—"

    def _max(lst: list) -> str:
        return f"{max(lst):.1f}" if lst else "—"

    for i, (ep_name, s) in enumerate(endpoint_stats.items()):
        r = header_row + 1 + i
        row_fill = _fill(COLOUR["alt_row"]) if i % 2 == 0 else _fill("FFFFFF")
        row_data = [
            ep_name, s["folder"], s["method"], s["url"],
            s["total"], s["passed"], s["failed"], s["warned"],
            _avg(s["ttfb_list"]), _avg(s["body_list"]), _avg(s["total_list"]),
            _min(s["total_list"]), _max(s["total_list"]),
        ]
        for c, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = _font()
            cell.fill = row_fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)

        # Colour the pass/fail/warn counts
        for col_offset, colour_key in ((6, "pass"), (7, "fail"), (8, "warn")):
            c = ws.cell(row=r, column=col_offset)
            bg, fg = _verdict_style(colour_key.upper())
            if c.value:
                c.fill = _fill(bg)
                c.font = _font(bold=True, colour=fg)

    _set_col_widths(ws, [30, 20, 10, 45, 14, 10, 10, 10, 15, 15, 15, 14, 14])


# ---------------------------------------------------------------------------
# Details sheet
# ---------------------------------------------------------------------------

DETAIL_HEADERS = [
    "Endpoint Name",
    "Folder",
    "Method",
    "URL",
    "Test Variant",
    "Variant Description",
    "Expected Status",
    "Actual Status",
    "Request Time / TTFB (ms)",   # time from send → headers received
    "Response Body Time (ms)",    # time to download body after headers
    "Total Time (ms)",            # TTFB + body download
    "LLM Verdict",
    "LLM Analysis",
    "Request Headers",
    "Request Body",
    "Response Body",
    "Error",
]

DETAIL_COL_WIDTHS = [28, 18, 10, 42, 22, 38, 15, 13, 22, 22, 16, 13, 55, 35, 35, 55, 30]


def _build_details(wb: Workbook, results: list[dict]) -> None:
    ws = wb.create_sheet("Details")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"  # freeze header row

    # Header row
    for c, h in enumerate(DETAIL_HEADERS, start=1):
        _header_cell(ws, 1, c, h)
    ws.row_dimensions[1].height = 22

    import json as _json

    for i, res in enumerate(results):
        r = i + 2
        row_fill = _fill(COLOUR["alt_row"]) if i % 2 == 0 else _fill("FFFFFF")

        analysis = res.get("analysis") or {}
        verdict = analysis.get("verdict", "FAIL" if res.get("error") else "")
        reason = analysis.get("reason", "")

        req_headers_str = _json.dumps(res.get("request_headers", {}), indent=None)

        ttfb    = res.get("request_time_ms", "")
        body_dl = res.get("response_time_ms", "")
        total   = res.get("total_time_ms", "")

        row_values = [
            res.get("endpoint_name", ""),
            res.get("folder_path", ""),
            res.get("method", ""),
            res.get("url", ""),
            res.get("variant_name", ""),
            res.get("variant_description", ""),
            res.get("expected_status", ""),
            res.get("status_code", ""),
            ttfb,
            body_dl,
            total,
            verdict,
            reason,
            req_headers_str,
            res.get("request_body", ""),
            res.get("response_body", "")[:2000],   # cap at 2000 chars
            res.get("error", "") or "",
        ]

        for c, val in enumerate(row_values, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = _font()
            cell.fill = row_fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=False)

        # Colour latency cells by threshold: >2000ms = red, >1000ms = yellow
        for lat_col, lat_val in ((9, ttfb), (10, body_dl), (11, total)):
            if isinstance(lat_val, (int, float)) and lat_val > 0:
                if lat_val > 2000:
                    ws.cell(row=r, column=lat_col).fill = _fill(COLOUR["fail_bg"])
                    ws.cell(row=r, column=lat_col).font = _font(bold=True, colour=COLOUR["fail_fg"])
                elif lat_val > 1000:
                    ws.cell(row=r, column=lat_col).fill = _fill(COLOUR["warn_bg"])
                    ws.cell(row=r, column=lat_col).font = _font(bold=True, colour=COLOUR["warn_fg"])

        # Colour the verdict cell (column 12)
        verdict_cell = ws.cell(row=r, column=12)
        if verdict:
            bg, fg = _verdict_style(verdict)
            verdict_cell.fill = _fill(bg)
            verdict_cell.font = _font(bold=True, colour=fg)

        # Wrap text for body/analysis columns
        for wrap_col in (13, 14, 15, 16, 17):
            ws.cell(row=r, column=wrap_col).alignment = Alignment(
                vertical="top", wrap_text=True
            )

        ws.row_dimensions[r].height = 60

    _set_col_widths(ws, DETAIL_COL_WIDTHS)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(DETAIL_HEADERS))}1"


# ---------------------------------------------------------------------------
# Latency sheet
# ---------------------------------------------------------------------------

LATENCY_HEADERS = [
    "Endpoint Name",
    "Method",
    "Test Variant",
    "Request Time / TTFB (ms)",
    "Response Body Time (ms)",
    "Total Time (ms)",
    "Status Code",
    "LLM Verdict",
    "Slow?",
]
LATENCY_COL_WIDTHS = [32, 10, 26, 24, 24, 18, 13, 13, 10]

# Thresholds for "Slow" flag
WARN_MS  = 1000
FAIL_MS  = 2000


def _build_latency(wb: Workbook, results: list[dict]) -> None:
    ws = wb.create_sheet("Latency")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    for c, h in enumerate(LATENCY_HEADERS, start=1):
        _header_cell(ws, 1, c, h)
    ws.row_dimensions[1].height = 22

    for i, res in enumerate(results):
        r = i + 2
        row_fill = _fill(COLOUR["alt_row"]) if i % 2 == 0 else _fill("FFFFFF")

        ttfb    = res.get("request_time_ms", 0.0)
        body_dl = res.get("response_time_ms", 0.0)
        total   = res.get("total_time_ms", 0.0)
        verdict = (res.get("analysis") or {}).get("verdict", "")

        if res.get("error"):
            slow_flag = "ERROR"
        elif isinstance(total, (int, float)) and total > FAIL_MS:
            slow_flag = "SLOW"
        elif isinstance(total, (int, float)) and total > WARN_MS:
            slow_flag = "WARN"
        else:
            slow_flag = ""

        row_values = [
            res.get("endpoint_name", ""),
            res.get("method", ""),
            res.get("variant_name", ""),
            ttfb,
            body_dl,
            total,
            res.get("status_code", ""),
            verdict,
            slow_flag,
        ]

        for c, val in enumerate(row_values, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = _font()
            cell.fill = row_fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

        # Colour latency value cells
        for lat_col, lat_val in ((4, ttfb), (5, body_dl), (6, total)):
            if isinstance(lat_val, (int, float)) and lat_val > 0:
                if lat_val > FAIL_MS:
                    ws.cell(row=r, column=lat_col).fill = _fill(COLOUR["fail_bg"])
                    ws.cell(row=r, column=lat_col).font = _font(bold=True, colour=COLOUR["fail_fg"])
                elif lat_val > WARN_MS:
                    ws.cell(row=r, column=lat_col).fill = _fill(COLOUR["warn_bg"])
                    ws.cell(row=r, column=lat_col).font = _font(bold=True, colour=COLOUR["warn_fg"])

        # Colour verdict cell (col 8)
        if verdict:
            bg, fg = _verdict_style(verdict)
            ws.cell(row=r, column=8).fill = _fill(bg)
            ws.cell(row=r, column=8).font = _font(bold=True, colour=fg)

        # Colour slow flag cell (col 9)
        if slow_flag == "SLOW":
            ws.cell(row=r, column=9).fill = _fill(COLOUR["fail_bg"])
            ws.cell(row=r, column=9).font = _font(bold=True, colour=COLOUR["fail_fg"])
        elif slow_flag == "WARN":
            ws.cell(row=r, column=9).fill = _fill(COLOUR["warn_bg"])
            ws.cell(row=r, column=9).font = _font(bold=True, colour=COLOUR["warn_fg"])
        elif slow_flag == "ERROR":
            ws.cell(row=r, column=9).fill = _fill(COLOUR["fail_bg"])
            ws.cell(row=r, column=9).font = _font(bold=True, colour=COLOUR["fail_fg"])

    _set_col_widths(ws, LATENCY_COL_WIDTHS)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(LATENCY_HEADERS))}1"


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def write_report(
    results: list[dict],
    output_path: str | Path,
    collection_name: str = "",
    started_at: datetime.datetime | None = None,
    finished_at: datetime.datetime | None = None,
) -> Path:
    """
    Write a two-sheet Excel report.

    Parameters
    ----------
    results : list[dict]
        Combined list where each item is a merged dict of:
            - run_test_case() output fields
            - ``analysis`` key → ollama_client.analyze_response() output
    output_path : str | Path
        Destination .xlsx file path.
    collection_name : str
        Human-readable label for the Summary sheet.
    started_at / finished_at : datetime
        Timestamps for the Summary sheet.

    Returns
    -------
    Path
        Resolved path to the written file.
    """
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    now = datetime.datetime.now()
    started_at = started_at or now
    finished_at = finished_at or now

    # Aggregate counts
    total = len(results)
    passed = sum(
        1 for r in results
        if (r.get("analysis") or {}).get("verdict", "").upper() == "PASS"
    )
    failed = sum(
        1 for r in results
        if (r.get("analysis") or {}).get("verdict", "").upper() == "FAIL"
    )
    warned = sum(
        1 for r in results
        if (r.get("analysis") or {}).get("verdict", "").upper() == "WARN"
    )
    errors = sum(1 for r in results if r.get("error"))

    run_meta = {
        "collection": collection_name,
        "started_at": started_at.strftime("%Y-%m-%d %H:%M:%S"),
        "finished_at": finished_at.strftime("%Y-%m-%d %H:%M:%S"),
        "total": total,
        "passed": passed,
        "failed": failed,
        "warned": warned,
        "errors": errors,
    }

    wb = Workbook()
    # Remove the default blank sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _build_summary(wb, results, run_meta)
    _build_details(wb, results)
    _build_latency(wb, results)

    wb.save(out)
    return out.resolve()
